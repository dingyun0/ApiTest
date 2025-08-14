# 氛围感空镜和人物情景空镜的dify工作流调用

from dotenv import load_dotenv
import os
import json
import sys
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from urllib.parse import urlparse
import urllib.request
import io
from PIL import Image as PILImage

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from DifyApi import DifyApi
load_dotenv()

dify_api=DifyApi(api_key=os.getenv("DIGITAL_PHOTOS_API_KEY"))
INPUT_EXCEL_PATH = "input/template.xlsx"  # 输入Excel文件路径
OUTPUT_EXCEL_PATH = "output/output.xlsx"  # 输出Excel文件路径
TEMP_EXCEL_PATH="temp/temp.xlsx"

file_path="/Users/donson/Documents/projects/ApiTest_vivian/ApiTest/text_to_video/dify/workflow/input/template.xlsx"
user="1234567890"

# 读取Excel文件
def read_excel(file_path):
    try:
        df = pd.read_excel(file_path)
        return df
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

# 下载图片到内存
def get_image_from_url(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        img_data = io.BytesIO(response.content)
        img = PILImage.open(img_data)
        return img
    except Exception as e:
        print(f"Error downloading image from {url}: {e}")
        return None

def run_dify(file_path,user):
    file_id=dify_api.upload_file(file_path,user)
    if file_id:
        print(f"文件上传成功，文件ID: {file_id}")
        inputs={
            "excel":
                {
                    "type": "document",
                    "transfer_method": "local_file",
                    "upload_file_id": file_id,
                }
        }
        response=dify_api.run_workflow(inputs,user)
        return response
    else:
        print("文件上传失败")

def download_excel(url,save_path):
    try:
        response=requests.get(url)
        response.raise_for_status()
        with open(save_path, 'wb') as f:
            f.write(response.content)
        return save_path
    except Exception as e:
        print(f"Error downloading Excel from {url}: {e}")
        return None

# 下载图片
def download_image(url, save_path):
    try:
        urllib.request.urlretrieve(url, save_path)
        return save_path
    except Exception as e:
        print(f"Error downloading image from {url}: {e}")
        return None

if __name__=="__main__":
    df=read_excel(INPUT_EXCEL_PATH)
    response=run_dify(file_path,user)
    files = response.get('data', {}).get('outputs', {}).get('files', [])
    excel_url = files[0].get('url')
    print(excel_url,'111111111')
    downloaded_excel = download_excel(excel_url, TEMP_EXCEL_PATH)
    processed_df = read_excel(TEMP_EXCEL_PATH)
    if 'image_url' not in processed_df.columns:
        print("Excel file must contain an 'image_url' column with image URLs")
    else:
        print('yes')

    # 创建新的Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入表头
    headers = processed_df.columns.tolist()
    for col_idx, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_idx)}1"] = header

    # 处理每一行数据并插入图片
    image_column_idx = headers.index('image_url') + 1 
    for row_idx, row in processed_df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws[f"{get_column_letter(col_idx)}{row_idx + 2}"]
            if headers[col_idx - 1] == 'image_url' and isinstance(value, str):
                # 下载图片到内存
                img = get_image_from_url(value)
                print(img,'2222222')
                if img:
                    # 保存图片到临时文件
                    temp_img_path = f"temp_image_{row_idx}.png"
                    img.save(temp_img_path)
                    # 插入图片到Excel
                    img_excel = Image(temp_img_path)
                    # 使用图片的原始尺寸（像素）
                    img_excel.width, img_excel.height = img.size[0], img.size[1]
                    # 动态调整单元格大小以适应图片
                    # 行高：1像素 ≈ 0.75点
                    ws.row_dimensions[row_idx + 2].height = img.size[1] * 0.75+1000
                    # 列宽：根据图片宽度动态调整（像素 / 5.5，实验优化）
                    current_col_width = ws.column_dimensions[get_column_letter(image_column_idx)].width or 10
                    new_col_width = img.size[0] / 5.5
                    if new_col_width > current_col_width:
                        ws.column_dimensions[get_column_letter(image_column_idx)].width = new_col_width
                    ws.add_image(img_excel, f"{get_column_letter(col_idx)}{row_idx + 2}")
                else:
                    cell.value = value  # 保留原始URL
            else:
                cell.value = value
            os.remove(temp_img_path)

    # 调整列宽以适应内容
    # 调整所有列宽以适应内容或图片
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        # 对于image_url列，确保宽度至少容纳最大图片
        if column == get_column_letter(image_column_idx):
            current_width = ws.column_dimensions[column].width or 10
            ws.column_dimensions[column].width = max(current_width, adjusted_width)
        else:
            ws.column_dimensions[column].width = max_length > 0 and adjusted_width or 10


    # 保存新的Excel文件
    try:
        wb.save(OUTPUT_EXCEL_PATH)
        print(f"Output Excel saved to {OUTPUT_EXCEL_PATH}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")



    








