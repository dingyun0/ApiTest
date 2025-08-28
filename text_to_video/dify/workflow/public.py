# 批量生成数字人图片的dify工作流调用后处理-封装版本

import os
import json
import sys
import io
import pandas as pd
import requests
import mimetypes
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from urllib.parse import urlparse
import tempfile
from PIL import Image as PILImage
from dotenv import load_dotenv

# 加载环境变量
# 方舟的apikey
# ARK_API_KEY='0892d12b-a090-4ce2-aeb4-4179b2e99ecb'
# # dify的base_url
# BASE_URL='http://113.45.180.205/v1'
# 数字人图片
# DIGITAL_PHOTOS_API_KEY='app-WpdKcHdS6lbQkVAytcVo42HY'
load_dotenv()

class DifyApi:
    def __init__(self, api_key=None, base_url=None):
        self.api_key = api_key
        self.base_url = os.getenv("BASE_URL")
        if not self.api_key or not self.base_url:
            raise ValueError("API_KEY或BASE_URL缺失")
    
    def upload_file(self, file_path, user):
        upload_url = f"{self.base_url}/files/upload"
        headers = {
            "Authorization": f"Bearer {self.api_key}",
        }
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type is None:
            mime_type = "application/octet-stream"
        try:
            print(f"正在上传文件: {file_path}...")
            with open(file_path, "rb") as f:
                file = {
                    "file": (os.path.basename(file_path), f, mime_type)
                }
                data = {
                    "user": user
                }
                response = requests.post(upload_url, headers=headers, files=file, data=data)
                if response.status_code in [200, 201]:
                    file_info = response.json()
                    print("文件上传成功！")
                    return file_info.get("id")
                else:
                    print(f"文件上传失败，状态码: {response.status_code}")
                    return None
        except FileNotFoundError:
            print(f"文件不存在: {file_path}")
            return None
        except Exception as e:
            print(f"上传文件时发生错误: {str(e)}")
            return None

    def run_workflow(self, inputs, user, response_mode="blocking"):
        workflow_url = f"{self.base_url}/workflows/run"
        headers = {
            'Authorization': f'Bearer {self.api_key}',
            'Content-Type': 'application/json'
        }
        
        data = {
            "inputs": inputs,
            "response_mode": response_mode,
            "user": user
        }

        try:
            print("正在运行工作流，请稍候...")
            response = requests.post(workflow_url, headers=headers, data=json.dumps(data))
            if response.status_code in [200, 201]:
                print("工作流运行成功！")
                return response.json()
            else:
                print(f"工作流运行失败，状态码: {response.status_code}")
                return {"status": "error", "message": f"Failed to execute workflow, status code: {response.status_code}"}
        except Exception as e:
            print(f"运行工作流时发生错误: {str(e)}")
            return {"status": "error", "message": str(e)}

# 初始化 Dify API
dify_api = DifyApi(api_key=os.getenv("DIGITAL_PHOTOS_API_KEY"))

# 文件路径
INPUT_EXCEL_PATH = os.getenv("INPUT_EXCEL_PATH")
OUTPUT_EXCEL_PATH = os.getenv("OUTPUT_EXCEL_PATH")
TEMP_EXCEL_PATH = os.getenv("TEMP_EXCEL_PATH")

# 用户 ID
USER_ID = "1234567890"


# 读取 Excel 文件
def read_excel(file_path):
    try:
        return pd.read_excel(file_path)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None


# 上传文件并运行 Dify 工作流
def run_dify(file_path, user_id):
    file_id = dify_api.upload_file(file_path, user_id)
    if file_id:
        print(f"文件上传成功，文件ID: {file_id}")
        inputs = {
            "excel": {
                "type": "document",
                "transfer_method": "local_file",
                "upload_file_id": file_id,
            }
        }
        response = dify_api.run_workflow(inputs, user_id)
        return response
    else:
        print("文件上传失败")
        return None


# 下载 Excel 文件
def download_excel(url, save_path):
    try:
        response = requests.get(url)
        response.raise_for_status()
        with open(save_path, 'wb') as f:
            f.write(response.content)
        return save_path
    except Exception as e:
        print(f"Error downloading Excel from {url}: {e}")
        return None

# 下载图片到内存
def get_image_from_url(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        img_data = io.BytesIO(response.content)
        return PILImage.open(img_data)
    except Exception as e:
        print(f"Error downloading image from {url}: {e}")
        return None



# 创建新的 Excel 文件并插入图片
def create_excel_with_images(processed_df, output_path):
    if 'image_url' not in processed_df.columns:
        print("Excel file must contain an 'image_url' column with image URLs")
        return

    # 创建新的 Excel 文件
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入表头
    headers = processed_df.columns.tolist()
    for col_idx, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_idx)}1"] = header

    # 处理每一行数据并插入图片
    image_column_idx = headers.index('image_url') + 1
    for row_idx, row in processed_df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws[f"{get_column_letter(col_idx)}{row_idx + 2}"]
            if headers[col_idx - 1] == 'image_url' and isinstance(value, str):
                # 下载图片到内存
                img = get_image_from_url(value)
                if img:
                    temp_img_path = f"temp_image_{row_idx}.png"
                    img.save(temp_img_path)
                    img_excel = Image(temp_img_path)
                    img_excel.width, img_excel.height = img.size[0], img.size[1]
                    ws.row_dimensions[row_idx + 2].height = img.size[1] * 0.75+1000
                    current_col_width = ws.column_dimensions[get_column_letter(image_column_idx)].width or 10
                    new_col_width = img.size[0] / 5.5
                    if new_col_width > current_col_width:
                        ws.column_dimensions[get_column_letter(image_column_idx)].width = new_col_width
                    ws.add_image(img_excel, f"{get_column_letter(col_idx)}{row_idx + 2}")
                else:
                    cell.value = value 
            else:
                cell.value = value

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # 保存新的 Excel 文件
    try:
        wb.save(output_path)
        print(f"Output Excel saved to {output_path}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")


# 主函数
if __name__ == "__main__":
    # 读取输入 Excel 文件
    df = read_excel(INPUT_EXCEL_PATH)
    if df is None:
        sys.exit("Failed to read input Excel file.")

    # 调用 Dify 工作流
    response = run_dify(INPUT_EXCEL_PATH, USER_ID)
    if not response:
        sys.exit("Failed to run Dify workflow.")

    # 下载处理后的 Excel 文件
    files = response.get('data', {}).get('outputs', {}).get('files', [])
    if not files:
        sys.exit("No files returned from Dify workflow.")
    excel_url = files[0].get('url')
    downloaded_excel = download_excel(excel_url, TEMP_EXCEL_PATH)
    if not downloaded_excel:
        sys.exit("Failed to download processed Excel file.")

    # 读取处理后的 Excel 文件
    processed_df = read_excel(TEMP_EXCEL_PATH)
    if processed_df is None:
        sys.exit("Failed to read processed Excel file.")

    # 创建新的 Excel 文件并插入图片
    create_excel_with_images(processed_df, OUTPUT_EXCEL_PATH)
